import sqlite3
from pathlib import Path
from datetime import datetime
from functools import wraps

import pandas as pd
from flask import Flask, render_template, request, redirect, url_for, session, send_file
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment


app = Flask(__name__)

app.secret_key = "clave_secreta_solo_para_pruebas"

ARCHIVO_BD = Path("sistema.db")
CARPETA_REPORTES = Path("reportes")
CARPETA_REPORTES.mkdir(exist_ok=True)


def conectar():
    return sqlite3.connect(ARCHIVO_BD)


def crear_tablas():
    conexion = conectar()
    cursor = conexion.cursor()

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS usuarios (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT NOT NULL,
            correo TEXT NOT NULL UNIQUE,
            password_hash TEXT NOT NULL,
            fecha_registro TEXT NOT NULL,
            rol TEXT NOT NULL DEFAULT 'usuario'
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS productos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT NOT NULL UNIQUE,
            precio REAL NOT NULL,
            stock INTEGER NOT NULL,
            fecha_registro TEXT NOT NULL
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS ventas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            usuario_id INTEGER NOT NULL,
            fecha TEXT NOT NULL,
            hora TEXT NOT NULL,
            producto TEXT NOT NULL,
            cantidad INTEGER NOT NULL,
            precio REAL NOT NULL,
            subtotal REAL NOT NULL,
            iva REAL NOT NULL,
            total REAL NOT NULL,
            FOREIGN KEY (usuario_id) REFERENCES usuarios(id)
        )
    """)

    conexion.commit()
    conexion.close()

    asegurar_columna_rol()


def asegurar_columna_rol():
    conexion = conectar()
    cursor = conexion.cursor()

    cursor.execute("PRAGMA table_info(usuarios)")
    columnas = cursor.fetchall()
    nombres_columnas = [columna[1] for columna in columnas]

    if "rol" not in nombres_columnas:
        cursor.execute("""
            ALTER TABLE usuarios
            ADD COLUMN rol TEXT NOT NULL DEFAULT 'usuario'
        """)

    conexion.commit()
    conexion.close()


def login_requerido(funcion):
    @wraps(funcion)
    def envoltura(*args, **kwargs):
        if "usuario_id" not in session:
            return redirect(url_for("login"))
        return funcion(*args, **kwargs)

    return envoltura


def admin_requerido(funcion):
    @wraps(funcion)
    def envoltura(*args, **kwargs):
        if "usuario_id" not in session:
            return redirect(url_for("login"))

        if session.get("usuario_rol") != "admin":
            return redirect(url_for("panel"))

        return funcion(*args, **kwargs)

    return envoltura


def buscar_usuario_por_correo(correo):
    conexion = conectar()
    cursor = conexion.cursor()

    cursor.execute("""
        SELECT id, nombre, correo, password_hash, rol
        FROM usuarios
        WHERE correo = ?
    """, (correo,))

    usuario = cursor.fetchone()
    conexion.close()

    return usuario


def buscar_usuario_por_id(usuario_id):
    conexion = conectar()
    cursor = conexion.cursor()

    cursor.execute("""
        SELECT id, nombre, correo, fecha_registro, rol
        FROM usuarios
        WHERE id = ?
    """, (usuario_id,))

    usuario = cursor.fetchone()
    conexion.close()

    return usuario


def obtener_productos_admin():
    conexion = conectar()
    cursor = conexion.cursor()

    cursor.execute("""
        SELECT id, nombre, precio, stock, fecha_registro
        FROM productos
        ORDER BY nombre
    """)

    productos = cursor.fetchall()
    conexion.close()

    return productos


def obtener_productos_disponibles():
    conexion = conectar()
    cursor = conexion.cursor()

    cursor.execute("""
        SELECT id, nombre, precio, stock
        FROM productos
        WHERE stock > 0
        ORDER BY nombre
    """)

    productos = cursor.fetchall()
    conexion.close()

    return productos


def agregar_producto_admin(nombre, precio, stock):
    nombre = nombre.strip().title()
    precio = float(precio)
    stock = int(stock)

    if not nombre:
        raise ValueError("Escribe el nombre del producto.")

    if precio <= 0:
        raise ValueError("El precio debe ser mayor a 0.")

    if stock < 0:
        raise ValueError("El stock no puede ser negativo.")

    conexion = conectar()
    cursor = conexion.cursor()

    cursor.execute("""
        SELECT id, stock
        FROM productos
        WHERE nombre = ?
    """, (nombre,))

    producto_existente = cursor.fetchone()

    if producto_existente:
        cursor.execute("""
            UPDATE productos
            SET precio = ?, stock = stock + ?
            WHERE nombre = ?
        """, (precio, stock, nombre))

        mensaje = "Producto actualizado correctamente. Se sumó el stock nuevo."
    else:
        fecha_registro = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        cursor.execute("""
            INSERT INTO productos (nombre, precio, stock, fecha_registro)
            VALUES (?, ?, ?, ?)
        """, (nombre, precio, stock, fecha_registro))

        mensaje = "Producto registrado correctamente."

    conexion.commit()
    conexion.close()

    return mensaje


def eliminar_producto_admin(producto_id):
    conexion = conectar()
    cursor = conexion.cursor()

    cursor.execute("""
        DELETE FROM productos
        WHERE id = ?
    """, (producto_id,))

    conexion.commit()
    conexion.close()


def registrar_venta(usuario_id, producto_id, cantidad):
    producto_id = int(producto_id)
    cantidad = int(cantidad)

    if cantidad <= 0:
        raise ValueError("La cantidad debe ser mayor a 0.")

    conexion = conectar()
    cursor = conexion.cursor()

    try:
        cursor.execute("""
            SELECT id, nombre, precio, stock
            FROM productos
            WHERE id = ?
        """, (producto_id,))

        producto = cursor.fetchone()

        if not producto:
            raise ValueError("Producto no encontrado.")

        id_producto, nombre_producto, precio, stock = producto

        if stock <= 0:
            raise ValueError("Este producto no tiene stock disponible.")

        if cantidad > stock:
            raise ValueError(f"No hay suficiente stock. Disponible: {stock}")

        ahora = datetime.now()
        fecha = ahora.strftime("%Y-%m-%d")
        hora = ahora.strftime("%H:%M:%S")

        subtotal = cantidad * precio
        iva = subtotal * 0.16
        total = subtotal + iva

        cursor.execute("""
            INSERT INTO ventas (
                usuario_id, fecha, hora, producto, cantidad, precio, subtotal, iva, total
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            usuario_id,
            fecha,
            hora,
            nombre_producto,
            cantidad,
            precio,
            subtotal,
            iva,
            total
        ))

        cursor.execute("""
            UPDATE productos
            SET stock = stock - ?
            WHERE id = ?
        """, (cantidad, producto_id))

        conexion.commit()

    except Exception:
        conexion.rollback()
        raise

    finally:
        conexion.close()


def obtener_ventas_usuario(usuario_id):
    conexion = conectar()
    cursor = conexion.cursor()

    cursor.execute("""
        SELECT id, fecha, hora, producto, cantidad, precio, subtotal, iva, total
        FROM ventas
        WHERE usuario_id = ?
        ORDER BY id DESC
    """, (usuario_id,))

    ventas = cursor.fetchall()
    conexion.close()

    return ventas


def obtener_resumen_usuario(usuario_id):
    conexion = conectar()
    cursor = conexion.cursor()

    cursor.execute("""
        SELECT
            COUNT(*),
            COALESCE(SUM(cantidad), 0),
            COALESCE(SUM(subtotal), 0),
            COALESCE(SUM(iva), 0),
            COALESCE(SUM(total), 0)
        FROM ventas
        WHERE usuario_id = ?
    """, (usuario_id,))

    resultado = cursor.fetchone()
    conexion.close()

    ventas_registradas, cantidad_total, subtotal, iva, total = resultado

    return {
        "ventas_registradas": ventas_registradas,
        "cantidad_total": cantidad_total,
        "subtotal": subtotal,
        "iva": iva,
        "total": total
    }


def eliminar_venta(usuario_id, venta_id):
    conexion = conectar()
    cursor = conexion.cursor()

    cursor.execute("""
        SELECT producto, cantidad
        FROM ventas
        WHERE id = ? AND usuario_id = ?
    """, (venta_id, usuario_id))

    venta = cursor.fetchone()

    if venta:
        producto, cantidad = venta

        cursor.execute("""
            DELETE FROM ventas
            WHERE id = ? AND usuario_id = ?
        """, (venta_id, usuario_id))

        cursor.execute("""
            UPDATE productos
            SET stock = stock + ?
            WHERE nombre = ?
        """, (cantidad, producto))

    conexion.commit()
    conexion.close()


def obtener_usuarios_admin():
    conexion = conectar()
    cursor = conexion.cursor()

    cursor.execute("""
        SELECT id, nombre, correo, rol, fecha_registro
        FROM usuarios
        ORDER BY id
    """)

    usuarios = cursor.fetchall()
    conexion.close()

    return usuarios


def obtener_ventas_admin():
    conexion = conectar()
    cursor = conexion.cursor()

    cursor.execute("""
        SELECT
            ventas.id,
            usuarios.nombre,
            usuarios.correo,
            ventas.fecha,
            ventas.hora,
            ventas.producto,
            ventas.cantidad,
            ventas.precio,
            ventas.subtotal,
            ventas.iva,
            ventas.total
        FROM ventas
        INNER JOIN usuarios
        ON ventas.usuario_id = usuarios.id
        ORDER BY ventas.id DESC
    """)

    ventas = cursor.fetchall()
    conexion.close()

    return ventas


def obtener_resumen_admin():
    conexion = conectar()
    cursor = conexion.cursor()

    cursor.execute("SELECT COUNT(*) FROM usuarios")
    total_usuarios = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*), COALESCE(SUM(stock), 0) FROM productos")
    total_productos, stock_total = cursor.fetchone()

    cursor.execute("""
        SELECT
            COUNT(*),
            COALESCE(SUM(cantidad), 0),
            COALESCE(SUM(iva), 0),
            COALESCE(SUM(total), 0)
        FROM ventas
    """)

    resultado = cursor.fetchone()
    conexion.close()

    total_ventas, cantidad, iva, total = resultado

    return {
        "usuarios": total_usuarios,
        "productos": total_productos,
        "stock": stock_total,
        "ventas": total_ventas,
        "cantidad": cantidad,
        "iva": iva,
        "total": total
    }


def crear_reporte_excel(usuario_id):
    conexion = conectar()

    ventas = pd.read_sql_query("""
        SELECT id, fecha, hora, producto, cantidad, precio, subtotal, iva, total
        FROM ventas
        WHERE usuario_id = ?
        ORDER BY id
    """, conexion, params=(usuario_id,))

    conexion.close()

    if ventas.empty:
        raise ValueError("No hay ventas registradas para generar el reporte.")

    resumen = pd.DataFrame({
        "Dato": [
            "Ventas registradas",
            "Cantidad total vendida",
            "Subtotal general",
            "IVA general",
            "Total vendido"
        ],
        "Valor": [
            len(ventas),
            ventas["cantidad"].sum(),
            ventas["subtotal"].sum(),
            ventas["iva"].sum(),
            ventas["total"].sum()
        ]
    })

    ventas_por_producto = (
        ventas.groupby("producto", as_index=False)
        .agg({
            "cantidad": "sum",
            "subtotal": "sum",
            "iva": "sum",
            "total": "sum"
        })
        .sort_values("total", ascending=False)
    )

    ventas_por_dia = (
        ventas.groupby("fecha", as_index=False)
        .agg({
            "cantidad": "sum",
            "subtotal": "sum",
            "iva": "sum",
            "total": "sum"
        })
        .sort_values("fecha")
    )

    fecha_archivo = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    ruta_reporte = CARPETA_REPORTES / f"reporte_usuario_{usuario_id}_{fecha_archivo}.xlsx"

    with pd.ExcelWriter(ruta_reporte, engine="openpyxl") as writer:
        resumen.to_excel(writer, sheet_name="Resumen", index=False)
        ventas.to_excel(writer, sheet_name="Ventas", index=False)
        ventas_por_producto.to_excel(writer, sheet_name="Ventas por producto", index=False)
        ventas_por_dia.to_excel(writer, sheet_name="Ventas por dia", index=False)

    dar_formato_excel(ruta_reporte)

    return ruta_reporte


def crear_reporte_general_excel():
    conexion = conectar()

    usuarios = pd.read_sql_query("""
        SELECT id, nombre, correo, rol, fecha_registro
        FROM usuarios
        ORDER BY id
    """, conexion)

    productos = pd.read_sql_query("""
        SELECT id, nombre, precio, stock, fecha_registro
        FROM productos
        ORDER BY nombre
    """, conexion)

    ventas = pd.read_sql_query("""
        SELECT
            ventas.id,
            usuarios.nombre AS usuario,
            usuarios.correo,
            ventas.fecha,
            ventas.hora,
            ventas.producto,
            ventas.cantidad,
            ventas.precio,
            ventas.subtotal,
            ventas.iva,
            ventas.total
        FROM ventas
        INNER JOIN usuarios
        ON ventas.usuario_id = usuarios.id
        ORDER BY ventas.id
    """, conexion)

    conexion.close()

    if ventas.empty:
        ventas_por_usuario = pd.DataFrame(columns=["usuario", "correo", "cantidad", "subtotal", "iva", "total"])
        ventas_por_producto = pd.DataFrame(columns=["producto", "cantidad", "subtotal", "iva", "total"])
        ventas_por_dia = pd.DataFrame(columns=["fecha", "cantidad", "subtotal", "iva", "total"])

        cantidad_total = 0
        subtotal_total = 0
        iva_total = 0
        total_general = 0
    else:
        ventas_por_usuario = (
            ventas.groupby(["usuario", "correo"], as_index=False)
            .agg({
                "cantidad": "sum",
                "subtotal": "sum",
                "iva": "sum",
                "total": "sum"
            })
            .sort_values("total", ascending=False)
        )

        ventas_por_producto = (
            ventas.groupby("producto", as_index=False)
            .agg({
                "cantidad": "sum",
                "subtotal": "sum",
                "iva": "sum",
                "total": "sum"
            })
            .sort_values("total", ascending=False)
        )

        ventas_por_dia = (
            ventas.groupby("fecha", as_index=False)
            .agg({
                "cantidad": "sum",
                "subtotal": "sum",
                "iva": "sum",
                "total": "sum"
            })
            .sort_values("fecha")
        )

        cantidad_total = ventas["cantidad"].sum()
        subtotal_total = ventas["subtotal"].sum()
        iva_total = ventas["iva"].sum()
        total_general = ventas["total"].sum()

    resumen = pd.DataFrame({
        "Dato": [
            "Usuarios registrados",
            "Productos registrados",
            "Stock total disponible",
            "Ventas registradas",
            "Cantidad total vendida",
            "Subtotal general",
            "IVA general",
            "Total general vendido"
        ],
        "Valor": [
            len(usuarios),
            len(productos),
            productos["stock"].sum() if not productos.empty else 0,
            len(ventas),
            cantidad_total,
            subtotal_total,
            iva_total,
            total_general
        ]
    })

    fecha_archivo = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    ruta_reporte = CARPETA_REPORTES / f"reporte_general_{fecha_archivo}.xlsx"

    with pd.ExcelWriter(ruta_reporte, engine="openpyxl") as writer:
        resumen.to_excel(writer, sheet_name="Resumen general", index=False)
        usuarios.to_excel(writer, sheet_name="Usuarios", index=False)
        productos.to_excel(writer, sheet_name="Inventario", index=False)
        ventas.to_excel(writer, sheet_name="Todas las ventas", index=False)
        ventas_por_usuario.to_excel(writer, sheet_name="Ventas por usuario", index=False)
        ventas_por_producto.to_excel(writer, sheet_name="Ventas por producto", index=False)
        ventas_por_dia.to_excel(writer, sheet_name="Ventas por dia", index=False)

    dar_formato_excel(ruta_reporte)

    return ruta_reporte


def dar_formato_excel(ruta_reporte):
    libro = load_workbook(ruta_reporte)

    for hoja in libro.worksheets:
        hoja.freeze_panes = "A2"

        for celda in hoja[1]:
            celda.font = Font(bold=True)
            celda.alignment = Alignment(horizontal="center")

        for columna in hoja.columns:
            ancho_maximo = 0
            letra = columna[0].column_letter

            for celda in columna:
                if celda.value is not None:
                    ancho_maximo = max(ancho_maximo, len(str(celda.value)))

                encabezado = hoja.cell(row=1, column=celda.column).value

                if encabezado in ["precio", "subtotal", "iva", "total", "Valor"]:
                    if isinstance(celda.value, (int, float)):
                        celda.number_format = '$#,##0.00'

            hoja.column_dimensions[letra].width = ancho_maximo + 8

    libro.save(ruta_reporte)


@app.route("/")
def inicio():
    return render_template("inicio.html")


@app.route("/registro", methods=["GET", "POST"])
def registro():
    error = None
    mensaje = None

    if request.method == "POST":
        nombre = request.form.get("nombre", "").strip().title()
        correo = request.form.get("correo", "").strip().lower()
        password = request.form.get("password", "")
        confirmar = request.form.get("confirmar", "")

        try:
            if not nombre or not correo or not password or not confirmar:
                raise ValueError("Completa todos los campos.")

            if "@" not in correo:
                raise ValueError("Escribe un correo válido.")

            if len(password) < 6:
                raise ValueError("La contraseña debe tener mínimo 6 caracteres.")

            if password != confirmar:
                raise ValueError("Las contraseñas no coinciden.")

            usuario_existente = buscar_usuario_por_correo(correo)

            if usuario_existente:
                raise ValueError("Ese correo ya está registrado.")

            password_hash = generate_password_hash(password)
            fecha_registro = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            conexion = conectar()
            cursor = conexion.cursor()

            cursor.execute("""
                INSERT INTO usuarios (nombre, correo, password_hash, fecha_registro, rol)
                VALUES (?, ?, ?, ?, 'usuario')
            """, (nombre, correo, password_hash, fecha_registro))

            conexion.commit()
            conexion.close()

            mensaje = "Usuario registrado correctamente. Ahora puedes iniciar sesión."

        except ValueError as problema:
            error = str(problema)

    return render_template("registro.html", error=error, mensaje=mensaje)


@app.route("/login", methods=["GET", "POST"])
def login():
    error = None

    if request.method == "POST":
        correo = request.form.get("correo", "").strip().lower()
        password = request.form.get("password", "")

        usuario = buscar_usuario_por_correo(correo)

        if not usuario:
            error = "Correo o contraseña incorrectos."
        else:
            usuario_id, nombre, correo_bd, password_hash, rol = usuario

            if check_password_hash(password_hash, password):
                session["usuario_id"] = usuario_id
                session["usuario_nombre"] = nombre
                session["usuario_rol"] = rol

                if rol == "admin":
                    return redirect(url_for("admin"))

                return redirect(url_for("panel"))

            error = "Correo o contraseña incorrectos."

    return render_template("login.html", error=error)


@app.route("/panel", methods=["GET", "POST"])
@login_requerido
def panel():
    error = None
    mensaje = None

    usuario_id = session["usuario_id"]
    usuario = buscar_usuario_por_id(usuario_id)

    if request.method == "POST":
        producto_id = request.form.get("producto_id", "")
        cantidad = request.form.get("cantidad", "")

        try:
            if not producto_id or not cantidad:
                raise ValueError("Selecciona producto y cantidad.")

            registrar_venta(usuario_id, producto_id, cantidad)
            mensaje = "Venta registrada correctamente. El stock fue actualizado."

        except ValueError as problema:
            error = str(problema)

    ventas = obtener_ventas_usuario(usuario_id)
    resumen = obtener_resumen_usuario(usuario_id)
    productos = obtener_productos_disponibles()

    return render_template(
        "panel.html",
        usuario=usuario,
        ventas=ventas,
        resumen=resumen,
        productos=productos,
        error=error,
        mensaje=mensaje
    )


@app.route("/admin")
@admin_requerido
def admin():
    usuario_id = session["usuario_id"]
    usuario = buscar_usuario_por_id(usuario_id)

    usuarios = obtener_usuarios_admin()
    ventas = obtener_ventas_admin()
    resumen_admin = obtener_resumen_admin()

    return render_template(
        "admin.html",
        usuario=usuario,
        usuarios=usuarios,
        ventas=ventas,
        resumen_admin=resumen_admin
    )


@app.route("/inventario", methods=["GET", "POST"])
@admin_requerido
def inventario():
    error = None
    mensaje = None

    usuario_id = session["usuario_id"]
    usuario = buscar_usuario_por_id(usuario_id)

    if request.method == "POST":
        nombre = request.form.get("nombre", "")
        precio = request.form.get("precio", "")
        stock = request.form.get("stock", "")

        try:
            if not nombre or not precio or not stock:
                raise ValueError("Completa nombre, precio y stock.")

            mensaje = agregar_producto_admin(nombre, precio, stock)

        except ValueError as problema:
            error = str(problema)

    productos = obtener_productos_admin()

    return render_template(
        "inventario.html",
        usuario=usuario,
        productos=productos,
        error=error,
        mensaje=mensaje
    )


@app.route("/eliminar-producto/<int:producto_id>")
@admin_requerido
def eliminar_producto(producto_id):
    eliminar_producto_admin(producto_id)
    return redirect(url_for("inventario"))


@app.route("/eliminar/<int:venta_id>")
@login_requerido
def eliminar(venta_id):
    usuario_id = session["usuario_id"]
    eliminar_venta(usuario_id, venta_id)

    return redirect(url_for("panel"))


@app.route("/descargar-reporte")
@login_requerido
def descargar_reporte():
    usuario_id = session["usuario_id"]

    try:
        ruta_reporte = crear_reporte_excel(usuario_id)

        return send_file(
            ruta_reporte,
            as_attachment=True,
            download_name=ruta_reporte.name
        )

    except Exception as error:
        usuario = buscar_usuario_por_id(usuario_id)
        ventas = obtener_ventas_usuario(usuario_id)
        resumen = obtener_resumen_usuario(usuario_id)
        productos = obtener_productos_disponibles()

        return render_template(
            "panel.html",
            usuario=usuario,
            ventas=ventas,
            resumen=resumen,
            productos=productos,
            error=str(error),
            mensaje=None
        )


@app.route("/descargar-reporte-general")
@admin_requerido
def descargar_reporte_general():
    try:
        ruta_reporte = crear_reporte_general_excel()

        return send_file(
            ruta_reporte,
            as_attachment=True,
            download_name=ruta_reporte.name
        )

    except Exception as error:
        usuario_id = session["usuario_id"]
        usuario = buscar_usuario_por_id(usuario_id)

        usuarios = obtener_usuarios_admin()
        ventas = obtener_ventas_admin()
        resumen_admin = obtener_resumen_admin()

        return render_template(
            "admin.html",
            usuario=usuario,
            usuarios=usuarios,
            ventas=ventas,
            resumen_admin=resumen_admin,
            error=str(error)
        )


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("inicio"))


if __name__ == "__main__":
    crear_tablas()
    app.run(debug=True)